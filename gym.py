import streamlit as st
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime, date, timedelta
from pathlib import Path

st.set_page_config(
    page_title="FitPro Manager",
    page_icon="🏋️",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={}
)

st.markdown("""
<style>
/* 1. Ẩn các thành phần thừa nhưng GIỮ LẠI NÚT SIDEBAR */
header[data-testid="stHeader"] { 
    background-color: rgba(0,0,0,0) !important;
    color: white !important;
}
/* Hiển thị lại nút đóng/mở Sidebar và làm nó nổi bật */
button[data-testid="sidebar-toggle"] {
    background-color: var(--primary) !important;
    color: white !important;
    display: block !important;
}

/* 2. Chặn việc ẩn Sidebar trên màn hình nhỏ */
[data-testid="stSidebar"] {
    display: block !important;
    visibility: visible !important;
}

#MainMenu, .stDeployButton, footer, [data-testid="stFooter"] { display:none !important; }
.block-container { padding-top: 2rem !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Nunito:wght@400;600;700;800&display=swap');

:root {
  --primary: #FF6B35;
  --secondary: #FFFFFF;
  --accent: #FFD700;
  --surface: #F8F9FA;
  --surface2: #E9ECEF;
  --text: #2D3436;
  --subtext: #636E72;
  --success: #27AE60;
  --warning: #F39C12;
  --danger: #D63031;
  --radius: 12px;
  --shadow: 0 4px 12px rgba(0,0,0,0.08);
}

html,body,[data-testid="stAppViewContainer"]{background:var(--secondary)!important;color:var(--text)!important;font-family:'Nunito',sans-serif!important;}
[data-testid="stSidebar"]{background:var(--surface)!important;border-right:2px solid var(--surface2);}
h1,h2,h3{font-family:'Bebas Neue',sans-serif!important;letter-spacing:1px;}
h1{color:var(--primary)!important;font-size:2.2rem!important;}
h2{color:var(--accent)!important;font-size:1.7rem!important;}
h3{color:var(--text)!important;font-size:1.2rem!important;}
.card{background:var(--surface);border-radius:var(--radius);padding:1.1rem 1.3rem;margin-bottom:.7rem;border-left:4px solid var(--primary);box-shadow:var(--shadow);transition:transform .15s;}
.card:hover{transform:translateY(-1px);}
.card-time{border-left-color:#A78BFA!important;}
.card-pt{border-left-color:#FF6B35!important;}
.card-ok{border-left-color:var(--success)!important;}
.card-warn{border-left-color:var(--warning)!important;}
.mbox{background:linear-gradient(135deg,var(--surface2),var(--surface));border-radius:var(--radius);padding:1rem 1.2rem;text-align:center;border:1px solid rgba(255,107,53,.3);box-shadow:var(--shadow);}
.mbox .val{font-family:'Bebas Neue';font-size:2rem;color:var(--primary);}
.mbox .lbl{font-size:.72rem;color:var(--subtext);text-transform:uppercase;letter-spacing:1px;}
.mbox .sub{font-size:.7rem;color:var(--success);margin-top:.2rem;}
.badge{display:inline-block;padding:.22rem .65rem;border-radius:20px;font-size:.72rem;font-weight:700;letter-spacing:.4px;text-transform:uppercase;}
.b-active{background:rgba(46,204,113,.18);color:#2ECC71;border:1px solid #2ECC71;}
.b-expired{background:rgba(231,76,60,.18);color:#E74C3C;border:1px solid #E74C3C;}
.b-pending{background:rgba(243,156,18,.18);color:#F39C12;border:1px solid #F39C12;}
.b-time{background:rgba(167,139,250,.18);color:#A78BFA;border:1px solid #A78BFA;}
.b-session{background:rgba(255,107,53,.18);color:#FF6B35;border:1px solid #FF6B35;}
.stButton>button{background:linear-gradient(135deg,var(--primary),#e0552a)!important;color:white!important;border:none!important;border-radius:8px!important;font-family:'Nunito'!important;font-weight:700!important;transition:opacity .2s,transform .1s!important;}
.stButton>button:hover{opacity:.88!important;transform:translateY(-1px)!important;}
.stTextInput input,.stTextArea textarea,.stSelectbox div[data-baseweb],.stDateInput input,.stNumberInput input{background:var(--surface)!important;color:var(--text)!important;border:1px solid var(--surface2)!important;border-radius:8px!important;}
.sidebar-logo{text-align:center;padding:1.1rem 0 .8rem;border-bottom:1px solid var(--surface2);margin-bottom:.8rem;}
.sidebar-logo .brand{font-family:'Bebas Neue';font-size:1.9rem;color:var(--primary);letter-spacing:2px;}
.sidebar-logo .tag{font-size:.65rem;color:var(--subtext);letter-spacing:2px;text-transform:uppercase;}
hr{border-color:var(--surface2)!important;margin:.8rem 0;}
.stTabs [data-baseweb="tab-list"]{background:var(--surface)!important;border-radius:10px;padding:.3rem;}
.stTabs [data-baseweb="tab"]{color:var(--subtext)!important;border-radius:8px;font-weight:600;}
.stTabs [aria-selected="true"]{background:var(--primary)!important;color:white!important;}
.confirm-box{background:linear-gradient(135deg,rgba(255,107,53,.12),rgba(255,107,53,.05));border:2px solid var(--primary);border-radius:var(--radius);padding:1.2rem 1.4rem;margin:.6rem 0;}
.confirm-box .confirm-title{font-family:'Bebas Neue';font-size:1.3rem;color:var(--primary);}
.prog-bar-wrap{flex:1;background:var(--surface2);border-radius:6px;height:8px;}
.prog-bar{background:var(--primary);height:8px;border-radius:6px;transition:width .5s;}
</style>
""", unsafe_allow_html=True)

LOGO_SVG = """<svg width="56" height="56" viewBox="0 0 60 60" fill="none" xmlns="http://www.w3.org/2000/svg">
  <rect width="60" height="60" rx="14" fill="#FF6B35"/>
  <rect x="8" y="27" width="44" height="6" rx="3" fill="white"/>
  <rect x="8" y="24" width="12" height="12" rx="3" fill="white"/>
  <rect x="40" y="24" width="12" height="12" rx="3" fill="white"/>
  <rect x="6" y="22" width="8" height="16" rx="4" fill="#FFD700"/>
  <rect x="46" y="22" width="8" height="16" rx="4" fill="#FFD700"/>
</svg>"""

@st.cache_resource
def init_firebase():
    if not firebase_admin._apps:
        if "firebase" in st.secrets:
            fb = st.secrets["firebase"]
            raw_key = fb["private_key"]
            clean_key = raw_key.replace("\\n", "\n").strip()
            cred = credentials.Certificate({
                "type": fb["type"],
                "project_id": fb["project_id"],
                "private_key_id": fb["private_key_id"],
                "private_key": clean_key,
                "client_email": fb["client_email"],
                "client_id": fb["client_id"],
                "auth_uri": fb["auth_uri"],
                "token_uri": fb["token_uri"],
                "auth_provider_x509_cert_url": fb["auth_provider_x509_cert_url"],
                "client_x509_cert_url": fb["client_x509_cert_url"],
            })
        else:
            key_path = Path(__file__).parent / "serviceAccountKey.json"
            if not key_path.exists(): return None
            cred = credentials.Certificate(str(key_path))
        firebase_admin.initialize_app(cred)
    return firestore.client()

db = init_firebase()

# ============================================================
# FIX: load_data_from_firebase — đọc đủ users + memberships
# ============================================================
def load_data_from_firebase():
    if not db:
        return
    try:
        # Load memberships
        memberships_ref = db.collection("memberships").stream()
        st.session_state.memberships = [{"id": doc.id, **doc.to_dict()} for doc in memberships_ref]

        # FIX: query role chữ thường "pt" (trước là "PT" — sai)
        pts_ref = db.collection("users").where("role", "==", "pt").stream()
        st.session_state.pts_list = [{"id": doc.id, **doc.to_dict()} for doc in pts_ref]

        # FIX: load customers — trước đây không có
        customers_ref = db.collection("users").where("role", "==", "customer").stream()
        st.session_state.customers_list = [{"id": doc.id, **doc.to_dict()} for doc in customers_ref]
    except Exception as e:
        st.warning(f"Lỗi tải dữ liệu Firebase: {e}")

MOCK_USERS = {
    "owner@fitpro.vn": {"password":"owner123","role":"owner","name":"Chủ Phòng Gym"},
    "pt1@fitpro.vn":   {"password":"123","role":"pt","name":"Nguyễn Văn A","pt_id":"pt1"},
    "pt2@fitpro.vn":   {"password":"pt456","role":"pt","name":"Trần Thị B","pt_id":"pt2"},
    "khach1@gmail.com":{"password":"kh123","role":"customer","name":"Lê Văn Cường","customer_id":"kh1"},
    "khach2@gmail.com":{"password":"kh456","role":"customer","name":"Nguyễn Thị Hà","customer_id":"kh2"},
}

MOCK_MEMBERSHIPS = [
    {"id":"m1","customer_id":"kh1","customer_name":"Lê Văn Cường",
     "type":"session","package_name":"Gói PT Premium",
     "pt_id":"pt1","pt_name":"PT Nguyễn Văn A",
     "sessions_total":36,"sessions_done":22,"start_date":"2025-01-01","status":"active","notes":"Mục tiêu giảm mỡ tăng cơ"},
    {"id":"m2","customer_id":"kh2","customer_name":"Nguyễn Thị Hà",
     "type":"time","package_name":"Gói Yoga 3 tháng",
     "pt_id":None,"pt_name":None,
     "start_date":"2025-02-01","end_date":"2025-05-01","status":"active","notes":"Yoga buổi sáng"},
    {"id":"m3","customer_id":"kh1","customer_name":"Lê Văn Cường",
     "type":"time","package_name":"Gói Gym 6 tháng",
     "pt_id":None,"pt_name":None,
     "start_date":"2025-01-15","end_date":"2025-07-15","status":"active","notes":"Tập tự do buổi tối"},
    {"id":"m4","customer_id":"kh2","customer_name":"Nguyễn Thị Hà",
     "type":"session","package_name":"Gói PT Starter",
     "pt_id":"pt2","pt_name":"PT Trần Thị B",
     "sessions_total":10,"sessions_done":8,"start_date":"2025-03-01","status":"active","notes":""},
]

MOCK_PENDING = [
    {"id":"p1","membership_id":"m1","customer_id":"kh1","customer_name":"Lê Văn Cường",
     "pt_id":"pt1","pt_name":"PT Nguyễn Văn A",
     "session_date":"2025-03-22","session_time":"09:00",
     "content":"Leg Day: Squat 4x10@60kg, Deadlift 4x8@80kg, Leg Press 3x12@100kg",
     "note":"Tăng tạ squat lần sau","status":"pending_customer"},
]

MOCK_SESSIONS = [
    {"id":"s1","membership_id":"m1","customer_id":"kh1","pt_id":"pt1",
     "session_date":"2025-03-20","session_time":"09:00",
     "content":"Upper Body: Bench 4x10@60kg, Row 4x10@50kg, OHP 3x10@40kg",
     "note":"Tốt, tăng tạ bench lần sau","weight":74.5,"confirmed":True},
    {"id":"s2","membership_id":"m1","customer_id":"kh1","pt_id":"pt1",
     "session_date":"2025-03-18","session_time":"09:00",
     "content":"Cardio 30ph + Core: Plank 3x60s, Ab wheel 3x15",
     "note":"","weight":75.0,"confirmed":True},
]

# Khởi tạo session_state
for k,v in {"logged_in":False,"user":None,"page":"dashboard","data_loaded":False}.items():
    if k not in st.session_state: st.session_state[k]=v
if "pending_list" not in st.session_state:
    st.session_state.pending_list=[p.copy() for p in MOCK_PENDING]
if "memberships" not in st.session_state:
    st.session_state.memberships=[m.copy() for m in MOCK_MEMBERSHIPS]
if "sessions_log" not in st.session_state:
    st.session_state.sessions_log=[s.copy() for s in MOCK_SESSIONS]
# FIX: khởi tạo customers_list và pts_list
if "customers_list" not in st.session_state:
    st.session_state.customers_list = []
if "pts_list" not in st.session_state:
    st.session_state.pts_list = []

def login(email, password):
    if db:
        for doc in db.collection("users").where("email", "==", email).limit(1).stream():
            u = doc.to_dict()
            if u.get("password") == password:
                st.session_state.logged_in = True
                st.session_state.user = {**u, "id": doc.id}
                # FIX: load dữ liệu Firebase ngay sau khi login thành công
                st.session_state.data_loaded = False
                return True

    if email in MOCK_USERS and MOCK_USERS[email]["password"] == password:
        u = MOCK_USERS[email].copy()
        u["email"] = email
        st.session_state.logged_in = True
        st.session_state.user = u
        return True

    return False

# FIX: Gọi load_data_from_firebase sau khi login, chỉ load 1 lần
if db and st.session_state.get("logged_in") and not st.session_state.get("data_loaded"):
    load_data_from_firebase()
    st.session_state.data_loaded = True

def days_left(end_date_str):
    try: return (date.fromisoformat(end_date_str)-date.today()).days
    except: return None

def fmt_date(d):
    try: return date.fromisoformat(d).strftime("%d/%m/%Y")
    except: return d

def status_color(d):
    if d is None: return "var(--subtext)"
    if d<0: return "var(--danger)"
    if d<=7: return "var(--warning)"
    return "var(--success)"

def get_my_memberships(uid): return [m for m in st.session_state.memberships if m.get("customer_id")==uid]
def get_pt_memberships(pid): return [m for m in st.session_state.memberships if m.get("pt_id")==pid]
def get_pending_for_customer(cid): return [p for p in st.session_state.pending_list if p.get("customer_id")==cid and p.get("status")=="pending_customer"]
def get_pending_for_pt(pid): return [p for p in st.session_state.pending_list if p.get("pt_id")==pid]
def sessions_remaining(m): return m.get("sessions_total",0)-m.get("sessions_done",0)

def render_membership_card(m,show_customer=False,show_pt=True):
    mtype=m.get("type","session")
    pkg=m.get("package_name","")
    if mtype=="time":
        dl=days_left(m.get("end_date",""))
        color=status_color(dl)
        status_text=f"Còn {dl} ngày" if dl is not None and dl>=0 else "Đã hết hạn"
        badge_cls="b-active" if (dl or 0)>0 else "b-expired"
        type_badge='<span class="badge b-time">⏱ Theo thời gian</span>'
        detail_html=f'<div style="font-size:.82rem;margin:.5rem 0">📅 <b>{fmt_date(m.get("start_date",""))}</b> → <b>{fmt_date(m.get("end_date",""))}</b> &nbsp;&nbsp;<span style="color:{color};font-weight:700">{status_text}</span></div>'
    else:
        done=m.get("sessions_done",0);total=m.get("sessions_total",0);left=total-done
        pct=int(done/total*100) if total else 0
        badge_cls="b-active" if left>0 else "b-expired"
        type_badge='<span class="badge b-session">🏋️ Theo buổi PT</span>'
        detail_html=f'<div style="font-size:.82rem;margin:.5rem 0 .3rem">💪 <b>{left} buổi còn lại</b> / {total} buổi &nbsp;<span style="color:var(--subtext)">({done} đã tập)</span></div><div class="prog-bar-wrap" style="margin:.3rem 0 .5rem"><div class="prog-bar" style="width:{pct}%"></div></div>'
    cline=f'<div style="font-size:.82rem;color:var(--subtext)">👤 {m.get("customer_name","")}</div>' if show_customer else ""
    pline=f'<div style="font-size:.82rem;color:var(--subtext)">💪 PT: {m.get("pt_name","")}</div>' if show_pt and m.get("pt_name") else ""
    nline=f'<div style="font-size:.8rem;color:var(--subtext);font-style:italic">📝 {m.get("notes","")}</div>' if m.get("notes") else ""
    badge_lbl="Còn hạn / Còn buổi" if badge_cls=="b-active" else "Hết hạn / Hết buổi"
    st.markdown(f'<div class="card {"card-time" if mtype=="time" else "card-pt"}"><div style="display:flex;justify-content:space-between;align-items:flex-start"><div><b style="font-size:1rem">{pkg}</b> <span class="badge {badge_cls}">{badge_lbl}</span></div><div>{type_badge}</div></div>{cline}{pline}{detail_html}{nline}</div>',unsafe_allow_html=True)

def render_sidebar():
    with st.sidebar:
        st.markdown(f'<div class="sidebar-logo">{LOGO_SVG}<div class="brand">FitPro</div><div class="tag">Gym Management System</div></div>',unsafe_allow_html=True)
        user=st.session_state.user;role=user.get("role")
        role_label={"owner":"👑 Chủ phòng gym","pt":"💪 Personal Trainer","customer":"🏃 Khách hàng"}.get(role,role)
        n_pending=0
        if role=="pt": n_pending=len(get_pending_for_pt(user.get("pt_id","")))
        elif role=="customer": n_pending=len(get_pending_for_customer(user.get("customer_id","")))
        st.markdown(f'<div class="card" style="margin-bottom:.8rem;border-left-color:var(--accent)"><div style="font-weight:700">{user.get("name","")}</div><div style="font-size:.78rem;color:var(--subtext)">{role_label}</div></div>',unsafe_allow_html=True)
        st.markdown("---")
        pages_map={"owner":[("📊","dashboard","Tổng quan"),("👥","customers","Khách hàng"),("🏋️","pts","Personal Trainer"),("📦","packages","Gói tập"),("📈","reports","Báo cáo")],"pt":[("📊","dashboard","Tổng quan"),("👥","my_customers","Khách của tôi"),("✍️","record_session","Ghi nhận buổi tập"),("📋","pt_sessions","Lịch sử buổi tập")],"customer":[("📊","dashboard","Tổng quan"),("🎫","my_memberships","Gói tập của tôi"),("✅","confirm_session","Xác nhận buổi tập"),("📖","my_sessions","Nhật ký tập"),("📈","progress","Kết quả")]}
        for icon,key,label in pages_map.get(role,[]):
            badge=""
            if key=="confirm_session" and n_pending:
                badge=f' 🔴{n_pending}'
            if st.button(f"{icon}  {label}{badge}",key=f"nav_{key}",use_container_width=True):
                st.session_state.page=key;st.rerun()
        st.markdown("---")
        # FIX: Nút reload dữ liệu Firebase
        if db and st.button("🔄  Tải lại dữ liệu",use_container_width=True):
            st.session_state.data_loaded = False
            load_data_from_firebase()
            st.session_state.data_loaded = True
            st.success("Đã tải lại!")
            st.rerun()
        if st.button("🚪  Đăng xuất",use_container_width=True):
            st.session_state.logged_in=False;st.session_state.user=None
            st.session_state.page="dashboard";st.session_state.data_loaded=False
            st.rerun()

def page_login():
    c1,c2,c3=st.columns([1,1.1,1])
    with c2:
        st.markdown(f'<div style="text-align:center;padding:2rem 0 1rem">{LOGO_SVG.replace("56","80")}<h1 style="margin:.5rem 0 .1rem">FitPro Manager</h1><p style="color:var(--subtext);font-size:.85rem">Hệ thống quản lý phòng tập chuyên nghiệp</p></div>',unsafe_allow_html=True)
        st.markdown("### 🔐 Đăng nhập")
        email=st.text_input("Email",placeholder="email@fitpro.vn")
        password=st.text_input("Mật khẩu",type="password",placeholder="••••••••")
        if st.button("Đăng nhập",use_container_width=True):
            if login(email,password): st.success("Đăng nhập thành công!"); st.rerun()
            else: st.error("Sai email hoặc mật khẩu!")
        if not db:
            st.markdown("""<div style="background:rgba(255,107,53,.08);border:1px solid rgba(255,107,53,.3);border-radius:10px;padding:1rem;margin-top:1rem;font-size:.8rem">
            <b>⚡ Demo mode</b><br><br>
            👑 owner@fitpro.vn / owner123<br>
            💪 pt1@fitpro.vn / pt123 &nbsp;|&nbsp; pt2@fitpro.vn / pt456<br>
            🏃 khach1@gmail.com / kh123 &nbsp;|&nbsp; khach2@gmail.com / kh456
            </div>""",unsafe_allow_html=True)

def page_dashboard():
    user=st.session_state.user;role=user.get("role")
    st.markdown(f"# 📊 Xin chào, {user.get('name','')}!")
    st.markdown(f"<span style='color:var(--subtext)'>{datetime.now().strftime('%A, %d/%m/%Y')}</span>",unsafe_allow_html=True)
    st.markdown("---")
    if role=="owner":
        all_m=st.session_state.memberships
        all_customers=st.session_state.get("customers_list",[])
        all_pts=st.session_state.get("pts_list",[])
        n_time=sum(1 for m in all_m if m["type"]=="time" and m["status"]=="active")
        n_sess=sum(1 for m in all_m if m["type"]=="session" and m["status"]=="active")
        # Tổng khách: ưu tiên customers_list Firebase, fallback từ memberships
        n_customers=len(all_customers) if all_customers else len({m["customer_id"] for m in all_m})
        n_pts=len(all_pts) if all_pts else 0
        c1,c2,c3,c4=st.columns(4)
        for lbl,val,sub,col in [
            ("Tổng số khách hàng",n_customers,"thành viên",c1),
            ("Tổng số PT",n_pts,"huấn luyện viên",c2),
            ("Tổng gói PT",n_sess,"đang active",c3),
            ("Tổng gói thời gian",n_time,"đang active",c4),
        ]:
            with col: st.markdown(f'<div class="mbox"><div class="val">{val}</div><div class="lbl">{lbl}</div><div class="sub">{sub}</div></div>',unsafe_allow_html=True)
        st.markdown("---")
        ca,cb=st.columns(2)
        with ca:
            st.markdown("## ⏱ Gói thời gian sắp hết")
            shown=False
            for m in st.session_state.memberships:
                if m["type"]=="time":
                    dl=days_left(m.get("end_date",""))
                    if dl is not None and 0<=dl<=14:
                        shown=True;color="var(--warning)" if dl>7 else "var(--danger)"
                        st.markdown(f'<div class="card card-warn"><b>{m["customer_name"]}</b> – {m["package_name"]}<span style="float:right;color:{color};font-weight:700">Còn {dl} ngày</span></div>',unsafe_allow_html=True)
            if not shown: st.info("Không có gói nào sắp hết hạn.")
        with cb:
            st.markdown("## 🏋️ Gói PT sắp hết buổi")
            shown=False
            for m in st.session_state.memberships:
                if m["type"]=="session":
                    left=sessions_remaining(m)
                    if 0<left<=5:
                        shown=True
                        st.markdown(f'<div class="card card-warn"><b>{m["customer_name"]}</b> – {m["package_name"]}<span style="float:right;color:var(--warning);font-weight:700">Còn {left} buổi</span></div>',unsafe_allow_html=True)
            if not shown: st.info("Không có gói nào sắp hết buổi.")

    elif role=="pt":
        my_m=get_pt_memberships(user.get("pt_id",""))
        pending=get_pending_for_pt(user.get("pt_id",""))
        active_customers={m["customer_id"] for m in my_m if m["status"]=="active"}
        total_sessions=sum(m.get("sessions_done",0) for m in my_m if m["type"]=="session")
        c1,c2,c3=st.columns(3)
        for lbl,val,col in [("Khách của tôi",len(active_customers),c1),("Chờ KH xác nhận",len(pending),c2),("Tổng buổi đã tập",total_sessions,c3)]:
            with col: st.markdown(f'<div class="mbox"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>',unsafe_allow_html=True)
        if pending:
            st.markdown("---"); st.markdown("## ⏳ Buổi chờ khách xác nhận")
            for p in pending:
                st.markdown(f'<div class="card card-warn"><b>{p["customer_name"]}</b> | {p["session_date"]} {p["session_time"]}<span class="badge b-pending" style="float:right">Chờ KH xác nhận</span><div style="font-size:.82rem;color:var(--subtext);margin-top:.4rem">📋 {p["content"]}</div></div>',unsafe_allow_html=True)
        st.markdown("---"); st.markdown("## Khách hàng & số buổi còn lại")
        for m in my_m:
            if m["type"]=="session": render_membership_card(m,show_customer=True,show_pt=False)

    else:
        cid=user.get("customer_id","");my_m=get_my_memberships(cid)
        pending=get_pending_for_customer(cid)
        time_pkgs=[m for m in my_m if m["type"]=="time" and m["status"]=="active"]
        session_pkgs=[m for m in my_m if m["type"]=="session" and m["status"]=="active"]
        total_left=sum(sessions_remaining(m) for m in session_pkgs)
        c1,c2,c3=st.columns(3)
        with c1: st.markdown(f'<div class="mbox"><div class="val">{len(time_pkgs)}</div><div class="lbl">Gói thời gian active</div></div>',unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="mbox"><div class="val">{total_left}</div><div class="lbl">Buổi PT còn lại</div></div>',unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="mbox"><div class="val">{len(pending)}</div><div class="lbl">Buổi chờ xác nhận</div></div>',unsafe_allow_html=True)
        if pending:
            st.markdown("---"); st.markdown("## 🔔 Buổi tập cần xác nhận")
            for p in pending:
                st.markdown(f'<div class="confirm-box"><div class="confirm-title">⚠️ PT ghi nhận buổi tập – cần xác nhận!</div><div style="margin:.5rem 0;font-size:.88rem">📅 <b>{p["session_date"]}</b> lúc <b>{p["session_time"]}</b> | 💪 {p["pt_name"]}</div><div style="font-size:.85rem;color:var(--subtext)">📋 {p["content"]}</div></div>',unsafe_allow_html=True)
            if st.button("✅ Xác nhận ngay"):
                st.session_state.page="confirm_session"; st.rerun()
        st.markdown("---"); st.markdown("## Gói tập của tôi")
        for m in my_m: render_membership_card(m,show_customer=False,show_pt=True)

def page_my_memberships():
    st.markdown("# 🎫 Gói tập của tôi"); st.markdown("---")
    cid=st.session_state.user.get("customer_id","")
    my_m=get_my_memberships(cid)
    if not my_m: st.info("Bạn chưa có gói tập nào."); return
    time_pkgs=[m for m in my_m if m["type"]=="time"]
    session_pkgs=[m for m in my_m if m["type"]=="session"]
    if time_pkgs:
        st.markdown("## ⏱ Gói tập theo thời gian")
        st.caption("Gym, yoga, zumba... – tập tự do trong thời hạn")
        for m in time_pkgs:
            dl=days_left(m.get("end_date",""));color=status_color(dl)
            render_membership_card(m,show_customer=False,show_pt=False)
            if dl is not None and dl>0:
                st.markdown(f'<div style="background:rgba(46,204,113,.06);border-radius:8px;padding:.6rem 1rem;margin-top:-.4rem;margin-bottom:.8rem;font-size:.85rem">⏰ Còn <b style="color:{color}">{dl} ngày</b> – hết hạn <b>{fmt_date(m.get("end_date",""))}</b></div>',unsafe_allow_html=True)
            elif dl is not None:
                st.markdown('<div style="background:rgba(231,76,60,.08);border-radius:8px;padding:.6rem 1rem;margin-top:-.4rem;margin-bottom:.8rem;font-size:.85rem;color:var(--danger)">⛔ Gói đã hết hạn</div>',unsafe_allow_html=True)
    if session_pkgs:
        st.markdown("---"); st.markdown("## 🏋️ Gói tập với PT (theo buổi)")
        st.caption("PT ghi nhận → Bạn xác nhận → Số buổi được trừ")
        for m in session_pkgs:
            done=m.get("sessions_done",0);total=m.get("sessions_total",0);left=total-done
            pct=int(done/total*100) if total else 0
            render_membership_card(m,show_customer=False,show_pt=True)
            st.markdown(f'<div style="background:rgba(255,107,53,.06);border-radius:8px;padding:.7rem 1rem;margin-top:-.4rem;margin-bottom:.8rem;font-size:.85rem">🏋️ <b style="color:var(--primary)">{left} buổi còn lại</b> | {done}/{total} đã sử dụng<div class="prog-bar-wrap" style="margin:.5rem 0 0"><div class="prog-bar" style="width:{pct}%"></div></div></div>',unsafe_allow_html=True)

def page_confirm_session():
    st.markdown("# ✅ Xác nhận buổi tập"); st.markdown("---")
    cid=st.session_state.user.get("customer_id","")
    pending=get_pending_for_customer(cid)
    if not pending:
        st.success("🎉 Không có buổi tập nào cần xác nhận!")
        st.markdown('<div class="card card-ok">Tất cả buổi tập đã được cập nhật!</div>',unsafe_allow_html=True)
        return
    st.markdown(f'<div style="background:rgba(243,156,18,.1);border:1px solid var(--warning);border-radius:10px;padding:.8rem 1rem;margin-bottom:1.2rem;font-size:.88rem">⚠️ Bạn có <b style="color:var(--warning)">{len(pending)} buổi tập</b> cần xác nhận. Sau khi xác nhận, số buổi sẽ tự động trừ.</div>',unsafe_allow_html=True)

    for idx, p in enumerate(pending):
        mid = p.get("membership_id", "")
        mem = next((m for m in st.session_state.memberships if m["id"] == mid), None)
        left = sessions_remaining(mem) if mem else "?"

        note_html = ""
        if p.get("note"):
            note_html = f'<div style="font-size:.82rem;color:var(--subtext)">📝 PT ghi chú: {p["note"]}</div>'

        st.markdown(f"""
            <div class="confirm-box">
                <div class="confirm-title">📋 Buổi tập #{idx+1}</div>
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:.4rem;margin:.6rem 0;font-size:.88rem">
                    <div>📅 Ngày: <b>{p["session_date"]}</b></div>
                    <div>🕐 Giờ: <b>{p["session_time"]}</b></div>
                    <div>💪 PT: <b>{p["pt_name"]}</b></div>
                    <div>🎫 Còn lại: <b style="color:var(--primary)">{left} buổi</b></div>
                </div>
                <div style="font-size:.88rem;margin:.4rem 0"><b>Nội dung:</b> {p["content"]}</div>
                {note_html}
            </div>
        """, unsafe_allow_html=True)

        col_ok, col_no, _ = st.columns([1, 1, 2])
        with col_ok:
            if st.button(f"✅ Xác nhận", key=f"confirm_{p['id']}", use_container_width=True):
                for pp in st.session_state.pending_list:
                    if pp["id"] == p["id"]: pp["status"] = "confirmed"
                for m in st.session_state.memberships:
                    if m["id"] == mid:
                        m["sessions_done"] = m.get("sessions_done", 0) + 1
                        if m["sessions_done"] >= m.get("sessions_total", 0): m["status"] = "completed"
                st.session_state.sessions_log.append({"id": f"s{len(st.session_state.sessions_log)+1}", "membership_id": mid, "customer_id": cid, "pt_id": p["pt_id"], "session_date": p["session_date"], "session_time": p["session_time"], "content": p["content"], "note": p.get("note", ""), "weight": 0, "confirmed": True})
                if db: db.collection("sessions").add({"membership_id": mid, "customer_id": cid, "session_date": p["session_date"], "content": p["content"], "confirmed": True, "confirmed_at": datetime.now().isoformat()})
                st.success("✅ Đã xác nhận! Số buổi đã được cập nhật."); st.rerun()
        with col_no:
            if st.button(f"❌ Từ chối", key=f"reject_{p['id']}", use_container_width=True):
                for pp in st.session_state.pending_list:
                    if pp["id"] == p["id"]: pp["status"] = "rejected"
                st.warning("Đã từ chối buổi tập."); st.rerun()
        st.markdown("<hr>", unsafe_allow_html=True)

def page_my_sessions():
    st.markdown("# 📖 Nhật ký tập"); st.markdown("---")
    cid = st.session_state.user.get("customer_id", "")

    logs = sorted(
        [s for s in st.session_state.sessions_log if s.get("customer_id") == cid and s.get("confirmed")],
        key=lambda x: x.get("session_date", ""),
        reverse=True
    )

    if not logs:
        st.info("Chưa có buổi tập nào được ghi nhận.")
        return

    for s in logs:
        s_note_html = ""
        if s.get("note"):
            s_note_html = f'<div style="font-size:.8rem;color:var(--subtext)">📝 {s["note"]}</div>'

        st.markdown(f"""
            <div class="card card-ok">
                <div style="display:flex;justify-content:space-between">
                    <b>{s.get("session_date","")} – {s.get("session_time","")}</b>
                    <span class="badge b-active">Đã xác nhận</span>
                </div>
                <div style="margin:.4rem 0;font-size:.88rem">📋 {s.get("content","")}</div>
                {s_note_html}
            </div>
        """, unsafe_allow_html=True)

def page_progress():
    st.markdown("# 📈 Kết quả & Tiến độ"); st.markdown("---")
    c1,c2=st.columns(2)
    with c1:
        st.markdown("### 📊 Chỉ số ban đầu")
        st.markdown('<div class="card"><div style="display:grid;grid-template-columns:1fr 1fr;gap:.5rem;font-size:.9rem"><div>⚖️ Cân nặng: <b>78 kg</b></div><div>📏 Chiều cao: <b>172 cm</b></div><div>🔥 % Mỡ: <b>28%</b></div><div>💪 % Cơ: <b>30%</b></div><div>📅 Ngày đo: <b>01/01/2025</b></div></div></div>',unsafe_allow_html=True)
    with c2:
        st.markdown("### 🏆 Chỉ số hiện tại")
        st.markdown('<div class="card card-ok"><div style="display:grid;grid-template-columns:1fr 1fr;gap:.5rem;font-size:.9rem"><div>⚖️ Cân nặng: <b style="color:#2ECC71">73 kg ▼5</b></div><div>📏 Chiều cao: <b>172 cm</b></div><div>🔥 % Mỡ: <b style="color:#2ECC71">22% ▼6</b></div><div>💪 % Cơ: <b style="color:#2ECC71">36% ▲6</b></div><div>📅 Ngày đo: <b>20/03/2025</b></div></div></div>',unsafe_allow_html=True)
    st.markdown("---"); st.markdown("### 📝 Nhận xét PT")
    st.text_area("Nhận xét kết thúc khóa",value="Học viên đã có tiến bộ rõ rệt sau 3 tháng. Cân nặng giảm 5kg, % mỡ giảm 6%, cơ bắp tăng đáng kể.",height=100,disabled=True)

def page_my_customers():
    st.markdown("# 👥 Khách hàng của tôi"); st.markdown("---")
    pt_id=st.session_state.user.get("pt_id","")
    my_m=get_pt_memberships(pt_id)
    if not my_m: st.info("Bạn chưa được phân công khách hàng nào."); return
    for m in my_m: render_membership_card(m,show_customer=True,show_pt=False)

def page_record_session():
    st.markdown("# ✍️ Ghi nhận buổi tập"); st.markdown("---")
    user=st.session_state.user;pt_id=user.get("pt_id","")
    my_m=[m for m in get_pt_memberships(pt_id) if m["type"]=="session" and m["status"]=="active" and sessions_remaining(m)>0]
    if not my_m: st.warning("Không có khách hàng nào còn buổi tập."); return
    st.markdown('<div style="background:rgba(255,107,53,.08);border:1px solid rgba(255,107,53,.3);border-radius:10px;padding:.8rem 1rem;margin-bottom:1.2rem;font-size:.88rem">📌 <b>Quy trình:</b> PT ghi nhận buổi tập → Khách hàng xác nhận → Số buổi tự động trừ cả 2 bên</div>',unsafe_allow_html=True)
    mem_options={f"{m['customer_name']} – {m['package_name']} (còn {sessions_remaining(m)} buổi)":m for m in my_m}
    sel_label=st.selectbox("Chọn khách hàng & gói tập *",list(mem_options.keys()))
    sel_mem=mem_options[sel_label];left=sessions_remaining(sel_mem)
    st.markdown(f'<div class="card" style="margin:.6rem 0;padding:.8rem 1rem"><div style="font-size:.85rem">👤 <b>{sel_mem["customer_name"]}</b> | 📦 {sel_mem["package_name"]} | <span style="color:var(--primary);font-weight:700">{left} buổi còn lại</span></div></div>',unsafe_allow_html=True)
    st.markdown("---")
    c1,c2=st.columns(2)
    with c1: s_date=st.date_input("Ngày tập *",value=date.today()); s_time=st.time_input("Giờ bắt đầu *")
    with c2: s_weight=st.number_input("Cân nặng khách (kg)",30.0,200.0,70.0); s_fat=st.number_input("% Mỡ cơ thể",0.0,60.0,20.0)
    s_content=st.text_area("Nội dung buổi tập *",placeholder="Squat 4x10@60kg, Deadlift 3x8@80kg...",height=110)
    s_note=st.text_area("Ghi chú / nhận xét PT",placeholder="Form tập, cần cải thiện...",height=80)
    st.markdown("---")
    if st.button("📤 Gửi yêu cầu xác nhận đến khách hàng",use_container_width=True):
        if not s_content.strip(): st.error("Vui lòng nhập nội dung buổi tập!")
        else:
            new_p={"id":f"p{len(st.session_state.pending_list)+1}","membership_id":sel_mem["id"],"customer_id":sel_mem["customer_id"],"customer_name":sel_mem["customer_name"],"pt_id":pt_id,"pt_name":user.get("name",""),"session_date":str(s_date),"session_time":str(s_time)[:5],"content":s_content.strip(),"note":s_note.strip(),"weight":s_weight,"fat":s_fat,"status":"pending_customer","created_at":datetime.now().isoformat()}
            st.session_state.pending_list.append(new_p)
            if db: db.collection("pending_sessions").add(new_p)
            st.success(f"✅ Đã gửi yêu cầu xác nhận đến {sel_mem['customer_name']}! Số buổi sẽ trừ sau khi khách xác nhận.")
            st.balloons()

def page_pt_sessions():
    st.markdown("# 📋 Lịch sử buổi tập"); st.markdown("---")
    pt_id=st.session_state.user.get("pt_id","")
    tab1,tab2=st.tabs(["✅ Đã xác nhận","⏳ Chờ xác nhận"])
    confirmed=[s for s in st.session_state.sessions_log if s.get("pt_id")==pt_id and s.get("confirmed")]
    pending=[p for p in st.session_state.pending_list if p.get("pt_id")==pt_id and p.get("status")=="pending_customer"]
    with tab1:
        if not confirmed: st.info("Chưa có buổi tập nào được xác nhận.")
        for s in sorted(confirmed,key=lambda x:x.get("session_date",""),reverse=True):
            mem=next((m for m in st.session_state.memberships if m["id"]==s.get("membership_id","")),None)
            cname=mem["customer_name"] if mem else s.get("customer_id","")
            st.markdown(f'<div class="card card-ok"><div style="display:flex;justify-content:space-between"><b>👤 {cname}</b><span class="badge b-active">Xác nhận</span></div><div style="font-size:.83rem;color:var(--subtext)">📅 {s.get("session_date","")} {s.get("session_time","")}</div><div style="font-size:.85rem;margin-top:.3rem">📋 {s.get("content","")}</div></div>',unsafe_allow_html=True)
    with tab2:
        if not pending: st.info("Không có buổi nào đang chờ xác nhận.")
        for p in pending:
            st.markdown(f'<div class="card card-warn"><div style="display:flex;justify-content:space-between"><b>👤 {p.get("customer_name","")}</b><span class="badge b-pending">Chờ KH xác nhận</span></div><div style="font-size:.83rem;color:var(--subtext)">📅 {p.get("session_date","")} {p.get("session_time","")}</div><div style="font-size:.85rem;margin-top:.3rem">📋 {p.get("content","")}</div></div>',unsafe_allow_html=True)

def page_customers_owner():
    st.markdown("# 👥 Quản lý Khách hàng"); st.markdown("---")

    all_memberships = st.session_state.memberships
    all_customers = st.session_state.get("customers_list", [])

    # Gộp tên từ cả 2 nguồn
    names_from_memberships = [m["customer_name"] for m in all_memberships]
    names_from_firebase = [c.get("name","") for c in all_customers]
    all_names = sorted(list(set(names_from_memberships + names_from_firebase)))

    tab1, tab2, tab3 = st.tabs(["👥 Danh sách khách hàng", "🎫 Giao gói tập", "➕ Thêm khách mới"])

    # ── TAB 1: Danh sách & chi tiết khách hàng ──
    with tab1:
        col_search, col_add = st.columns([3,1])
        with col_search:
            selected_customer = st.selectbox("🔍 Chọn khách hàng để xem chi tiết:", ["-- Chọn khách hàng --"] + all_names)

        if selected_customer != "-- Chọn khách hàng --":
            # Tìm thông tin cá nhân từ customers_list
            cust_info = next((c for c in all_customers if c.get("name","") == selected_customer), None)
            cust_memberships = [m for m in all_memberships if m.get("customer_name","") == selected_customer]

            st.markdown(f"### 👤 {selected_customer}")

            # --- Thông tin cá nhân ---
            with st.expander("📋 Thông tin cá nhân", expanded=True):
                if cust_info:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(f"""
                        <div class="card">
                            <div style="font-size:.9rem;line-height:1.9">
                                📱 <b>SĐT:</b> {cust_info.get("phone","—")}<br>
                                🚻 <b>Giới tính:</b> {cust_info.get("gender","—")}<br>
                                🎂 <b>Ngày sinh:</b> {fmt_date(cust_info.get("dob",""))}<br>
                                📅 <b>Ngày đăng ký:</b> {fmt_date(cust_info.get("created_at","")[:10] if cust_info.get("created_at") else "")}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    with c2:
                        st.markdown(f"""
                        <div class="card">
                            <div style="font-size:.9rem;line-height:1.9">
                                📧 <b>Email:</b> {cust_info.get("email","—")}<br>
                                🔑 <b>Mật khẩu:</b> {"•"*len(cust_info.get("password","")) if cust_info.get("password") else "—"}<br>
                                📝 <b>Ghi chú:</b> {cust_info.get("note","—")}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("Không tìm thấy thông tin cá nhân trong hệ thống (khách từ gói tập).")

            # --- Gói tập hiện có ---
            if cust_memberships:
                st.markdown("#### 🎫 Gói tập đang có")
                for m in cust_memberships:
                    render_membership_card(m, show_customer=False, show_pt=True)
            else:
                # Khách mới chưa có gói
                st.markdown("""
                <div class="card" style="border-left-color:#A78BFA;background:rgba(167,139,250,.08)">
                    <span class="badge b-pending">🆕 Chưa có gói tập</span>
                    <div style="margin-top:.6rem;font-size:.88rem;color:var(--subtext)">Khách hàng này chưa được giao gói tập nào. Hãy chọn gói tập phù hợp bên dưới.</div>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("**🎯 Giao gói tập nhanh:**")
                pts_from_firebase = st.session_state.get("pts_list", [])
                pt_options_map = {}
                for pt in pts_from_firebase:
                    label = f"{pt.get('name','')} ({pt.get('specialty','')})"
                    pt_options_map[label] = (pt.get("id",""), pt.get("name",""))
                if not pt_options_map:
                    pt_options_map = {"PT Nguyễn Văn A (Cardio)": ("pt1","PT Nguyễn Văn A"), "PT Trần Thị B (Yoga)": ("pt2","PT Trần Thị B")}

                qc1, qc2 = st.columns(2)
                with qc1:
                    q_pkg = st.text_input("Tên gói tập *", key=f"q_pkg_{selected_customer}")
                    q_type = st.selectbox("Loại gói *", ["⏱ Theo thời gian", "🏋️ Theo buổi PT"], key=f"q_type_{selected_customer}")
                with qc2:
                    q_start = st.date_input("Ngày bắt đầu", value=date.today(), key=f"q_start_{selected_customer}")
                    if "thời gian" in q_type:
                        q_dur = st.selectbox("Thời hạn", ["1 tháng (30 ngày)","3 tháng (90 ngày)","6 tháng (180 ngày)"], key=f"q_dur_{selected_customer}")
                        dur_map2={"1 tháng (30 ngày)":30,"3 tháng (90 ngày)":90,"6 tháng (180 ngày)":180}
                        q_end = q_start + timedelta(days=dur_map2[q_dur])
                    else:
                        q_sess = st.number_input("Số buổi", 1, 500, 10, key=f"q_sess_{selected_customer}")
                        q_pt = st.selectbox("Chọn PT *", list(pt_options_map.keys()), key=f"q_pt_{selected_customer}")
                        q_pt_id, q_pt_name = pt_options_map[q_pt]

                if st.button("💾 Giao gói tập ngay", key=f"q_assign_{selected_customer}", use_container_width=True):
                    if q_pkg:
                        cust_id = cust_info.get("id", f"kh_{selected_customer}") if cust_info else f"kh_{selected_customer}"
                        new_m = {"id": f"m{len(st.session_state.memberships)+1}", "customer_id": cust_id, "customer_name": selected_customer, "package_name": q_pkg, "start_date": str(q_start), "status": "active", "notes": ""}
                        if "thời gian" in q_type:
                            new_m.update({"type":"time","end_date":str(q_end),"pt_id":None,"pt_name":None})
                        else:
                            new_m.update({"type":"session","sessions_total":q_sess,"sessions_done":0,"pt_id":q_pt_id,"pt_name":q_pt_name})
                        st.session_state.memberships.append(new_m)
                        if db: db.collection("memberships").add(new_m)
                        st.success(f"✅ Đã giao gói '{q_pkg}' cho {selected_customer}!"); st.rerun()
                    else:
                        st.error("Vui lòng nhập tên gói tập!")

            # --- Xóa khách hàng ---
            st.markdown("---")
            with st.expander("🗑️ Xóa khách hàng này", expanded=False):
                st.markdown(f'<div class="confirm-box"><div class="confirm-title">⚠️ Xác nhận xóa khách hàng</div><div style="font-size:.88rem;margin-top:.5rem">Bạn sắp xóa <b>{selected_customer}</b> khỏi hệ thống. Hành động này không thể hoàn tác!</div></div>', unsafe_allow_html=True)
                confirm_name = st.text_input("Nhập lại tên khách hàng để xác nhận xóa:", key=f"del_confirm_{selected_customer}")
                if st.button("🗑️ Xác nhận xóa", key=f"del_btn_{selected_customer}", type="primary"):
                    if confirm_name == selected_customer:
                        # Xóa khỏi customers_list
                        st.session_state.customers_list = [c for c in st.session_state.customers_list if c.get("name","") != selected_customer]
                        # Xóa gói tập liên quan
                        st.session_state.memberships = [m for m in st.session_state.memberships if m.get("customer_name","") != selected_customer]
                        if db and cust_info and cust_info.get("id"):
                            db.collection("users").document(cust_info["id"]).delete()
                        st.success(f"✅ Đã xóa khách hàng {selected_customer}!"); st.rerun()
                    else:
                        st.error("Tên xác nhận không khớp! Vui lòng nhập đúng tên khách hàng.")
        else:
            # Hiển thị tất cả khách hàng dạng danh sách
            st.markdown(f"#### 📋 Tổng số: {len(all_names)} khách hàng")
            members_with_pkg = {m["customer_name"] for m in all_memberships}
            for name in all_names:
                cust_info2 = next((c for c in all_customers if c.get("name","") == name), None)
                has_pkg = name in members_with_pkg
                pkg_count = len([m for m in all_memberships if m["customer_name"] == name])
                badge = f'<span class="badge b-active">{pkg_count} gói tập</span>' if has_pkg else '<span class="badge b-pending">Chưa có gói</span>'
                phone = cust_info2.get("phone","—") if cust_info2 else "—"
                gender = cust_info2.get("gender","—") if cust_info2 else "—"
                st.markdown(f'''
                    <div class="card">
                        <div style="display:flex;justify-content:space-between;align-items:center">
                            <b>👤 {name}</b>{badge}
                        </div>
                        <div style="font-size:.82rem;color:var(--subtext);margin-top:.4rem">
                            📱 {phone} &nbsp;|&nbsp; 🚻 {gender}
                        </div>
                    </div>
                ''', unsafe_allow_html=True)

    # ── TAB 2: Giao gói tập ──
    with tab2:
        st.markdown("### 🎫 Giao gói tập cho khách hàng")
        st.markdown('<div style="background:rgba(167,139,250,.08);border:1px solid rgba(167,139,250,.3);border-radius:10px;padding:.8rem 1rem;margin-bottom:1rem;font-size:.85rem">📌 <b>⏱ Theo thời gian:</b> Gym, Yoga... → cộng thêm ngày từ ngày mua<br>📌 <b>🏋️ Theo buổi PT:</b> Tập 1-1 với PT → PT ghi nhận, KH xác nhận từng buổi</div>',unsafe_allow_html=True)
        c1,c2=st.columns(2)
        with c1:
            g_customer_sel = st.selectbox("Chọn khách hàng *", ["-- Chọn --"] + all_names, key="g_cust_sel")
            g_customer = g_customer_sel if g_customer_sel != "-- Chọn --" else ""
            g_pkg=st.text_input("Tên gói tập *",placeholder="Gói Yoga 3 tháng")
        with c2:
            g_type=st.selectbox("Loại gói *",["⏱ Theo thời gian","🏋️ Theo buổi PT"])
            g_start=st.date_input("Ngày bắt đầu",value=date.today())

        pts_from_firebase = st.session_state.get("pts_list", [])
        pt_options_map = {}
        for pt in pts_from_firebase:
            label = f"{pt.get('name','')} ({pt.get('id','')[:6]})"
            pt_options_map[label] = (pt.get("id",""), pt.get("name",""))
        if not pt_options_map:
            pt_options_map = {"PT Nguyễn Văn A (pt1)": ("pt1","PT Nguyễn Văn A"), "PT Trần Thị B (pt2)": ("pt2","PT Trần Thị B")}

        if "thời gian" in g_type:
            g_duration=st.selectbox("Thời hạn",["1 tháng (30 ngày)","3 tháng (90 ngày)","6 tháng (180 ngày)","1 năm (365 ngày)"])
            dur_map={"1 tháng (30 ngày)":30,"3 tháng (90 ngày)":90,"6 tháng (180 ngày)":180,"1 năm (365 ngày)":365}
            dur_days=dur_map[g_duration];g_end=g_start+timedelta(days=dur_days)
            st.markdown(f'<div class="card card-time" style="padding:.7rem 1rem;font-size:.85rem">⏰ Ngày kết thúc: <b>{fmt_date(str(g_end))}</b> (sau {dur_days} ngày)</div>',unsafe_allow_html=True)
        else:
            g_sessions=st.number_input("Số buổi tập *",1,500,12)
            g_pt=st.selectbox("Phân công PT *", list(pt_options_map.keys()))
            pt_id_sel, pt_name_sel = pt_options_map[g_pt]

        g_note=st.text_input("Ghi chú")
        if st.button("💾 Giao gói tập",use_container_width=True):
            if g_customer and g_pkg:
                cust_obj = next((c for c in all_customers if c.get("name","") == g_customer), None)
                g_cid = cust_obj.get("id", f"kh{len(st.session_state.memberships)+1}") if cust_obj else f"kh{len(st.session_state.memberships)+1}"
                new_m={"id":f"m{len(st.session_state.memberships)+1}","customer_id":g_cid,"customer_name":g_customer,"package_name":g_pkg,"start_date":str(g_start),"status":"active","notes":g_note}
                if "thời gian" in g_type: new_m.update({"type":"time","end_date":str(g_end),"pt_id":None,"pt_name":None})
                else: new_m.update({"type":"session","sessions_total":g_sessions,"sessions_done":0,"pt_id":pt_id_sel,"pt_name":pt_name_sel})
                st.session_state.memberships.append(new_m)
                if db: db.collection("memberships").add(new_m)
                st.success(f"✅ Đã giao gói '{g_pkg}' cho {g_customer}!"); st.rerun()
            else: st.error("Vui lòng chọn khách hàng và nhập tên gói tập!")

    # ── TAB 3: Thêm khách hàng mới ──
    with tab3:
        st.markdown("### ➕ Thêm khách hàng mới")
        c1,c2=st.columns(2)
        with c1:
            n_name=st.text_input("Họ và tên *");n_email=st.text_input("Email *")
            n_phone=st.text_input("Số điện thoại");n_pwd=st.text_input("Mật khẩu *",type="password")
        with c2:
            n_gender=st.selectbox("Giới tính",["Nam","Nữ","Khác"])
            n_dob=st.date_input("Ngày sinh",value=date(1995,1,1))
            n_note=st.text_area("Ghi chú ban đầu")
        c1b,c2b,c3b,c4b=st.columns(4)
        with c1b: n_w=st.number_input("Cân nặng (kg)",30.0,200.0,65.0)
        with c2b: n_h=st.number_input("Chiều cao (cm)",100.0,220.0,165.0)
        with c3b: n_f=st.number_input("% Mỡ",0.0,60.0,20.0)
        with c4b: n_m=st.number_input("% Cơ",0.0,60.0,35.0)
        if st.button("💾 Thêm khách hàng",use_container_width=True):
            if n_name and n_email and n_pwd:
                if db:
                    doc_ref = db.collection("users").add({
                        "name":n_name,"email":n_email,"phone":n_phone,"password":n_pwd,
                        "role":"customer","gender":n_gender,"dob":str(n_dob),
                        "initial":{"weight":n_w,"height":n_h,"fat":n_f,"muscle":n_m},
                        "note":n_note,"created_at":datetime.now().isoformat()
                    })
                st.session_state.customers_list.append({
                    "name":n_name,"email":n_email,"phone":n_phone,"password":n_pwd,
                    "role":"customer","gender":n_gender,"dob":str(n_dob),
                    "note":n_note,"created_at":datetime.now().isoformat()
                })
                st.success(f"✅ Đã thêm khách hàng {n_name}!"); st.rerun()
            else: st.error("Điền đầy đủ thông tin bắt buộc!")

def page_packages():
    st.markdown("# 📦 Quản lý Gói tập"); st.markdown("---")
    st.markdown('<div style="display:grid;grid-template-columns:1fr 1fr;gap:1rem;margin-bottom:1.5rem"><div class="card card-time" style="padding:1rem 1.2rem"><div style="font-family:\'Bebas Neue\';font-size:1.2rem;color:#A78BFA">⏱ GÓI THEO THỜI GIAN</div><div style="font-size:.85rem;margin:.4rem 0;color:var(--subtext)">Gym, yoga, zumba...<br>Tập tự do trong thời hạn<br>Hệ thống tính ngày kết thúc tự động</div></div><div class="card card-pt" style="padding:1rem 1.2rem"><div style="font-family:\'Bebas Neue\';font-size:1.2rem;color:#FF6B35">🏋️ GÓI THEO BUỔI PT</div><div style="font-size:.85rem;margin:.4rem 0;color:var(--subtext)">Tập 1-1 với Personal Trainer<br>PT ghi nhận → KH xác nhận → trừ buổi<br>Cả PT và KH thấy số buổi còn lại</div></div></div>',unsafe_allow_html=True)

    all_m = st.session_state.memberships

    # Custom packages trong session (tách khỏi memberships)
    if "custom_packages" not in st.session_state:
        st.session_state.custom_packages = [
            {"id":"cp1","name":"Gói Gym Basic 1T","type":"time","price":800000,"duration":30,"note":"Tập tự do"},
            {"id":"cp2","name":"Gói Yoga 3T","type":"time","price":2200000,"duration":90,"note":"Yoga & Zumba"},
            {"id":"cp3","name":"Gói All-in-One 6T","type":"time","price":3800000,"duration":180,"note":"Tất cả dịch vụ"},
            {"id":"cp4","name":"Gói PT Starter","type":"session","price":1500000,"sessions":10,"note":""},
            {"id":"cp5","name":"Gói PT Premium","type":"session","price":3800000,"sessions":36,"note":""},
            {"id":"cp6","name":"Gói PT VIP","type":"session","price":6500000,"sessions":72,"note":""},
        ]

    tab1, tab2, tab3 = st.tabs(["📋 Gói hiện có & Cảnh báo", "➕ Tạo gói mới", "🗑️ Xóa gói"])

    with tab1:
        # Cảnh báo gói sắp hết
        warn_time = [m for m in all_m if m["type"]=="time" and days_left(m.get("end_date","")) is not None and 0 <= days_left(m.get("end_date","")) <= 14]
        warn_sess = [m for m in all_m if m["type"]=="session" and 0 < sessions_remaining(m) <= 5]

        if warn_time or warn_sess:
            st.markdown("### ⚠️ Cảnh báo")
            wc1, wc2 = st.columns(2)
            with wc1:
                st.markdown("**⏱ Gói thời gian sắp hết hiệu lực:**")
                if warn_time:
                    for m in warn_time:
                        dl = days_left(m.get("end_date",""))
                        color = "var(--danger)" if dl <= 7 else "var(--warning)"
                        # Hiển thị tên KH nếu click vào gói
                        st.markdown(f'<div class="card card-warn"><b>{m["package_name"]}</b><br><span style="font-size:.82rem;color:var(--subtext)">👤 {m["customer_name"]}</span><span style="float:right;color:{color};font-weight:700">Còn {dl} ngày</span></div>', unsafe_allow_html=True)
                else:
                    st.info("Không có cảnh báo.")
            with wc2:
                st.markdown("**🏋️ Gói PT sắp hết buổi:**")
                if warn_sess:
                    for m in warn_sess:
                        left = sessions_remaining(m)
                        st.markdown(f'<div class="card card-warn"><b>{m["package_name"]}</b><br><span style="font-size:.82rem;color:var(--subtext)">👤 {m["customer_name"]} | 💪 {m.get("pt_name","")}</span><span style="float:right;color:var(--warning);font-weight:700">Còn {left} buổi</span></div>', unsafe_allow_html=True)
                else:
                    st.info("Không có cảnh báo.")
            st.markdown("---")

        # Danh sách gói
        st.markdown("### ⏱ Gói theo thời gian")
        time_pkgs = [p for p in st.session_state.custom_packages if p["type"]=="time"]
        if time_pkgs:
            cols = st.columns(min(len(time_pkgs), 3))
            for i, pkg in enumerate(time_pkgs):
                # Đếm KH đang dùng gói này
                users_on_pkg = [m["customer_name"] for m in all_m if m.get("package_name","") == pkg["name"] and m["status"]=="active"]
                with cols[i % 3]:
                    users_html = f'<div style="font-size:.75rem;color:var(--success);margin-top:.3rem">👥 {", ".join(users_on_pkg)}</div>' if users_on_pkg else ""
                    st.markdown(f'<div class="card card-time"><div style="font-family:\'Bebas Neue\';font-size:1.1rem;color:#A78BFA">{pkg["name"]}</div><div style="font-size:1.4rem;font-weight:800;margin:.3rem 0">{pkg["price"]:,.0f}<span style="font-size:.75rem"> đ</span></div><div style="font-size:.82rem;color:var(--subtext)">🗓 {pkg["duration"]} ngày</div>{users_html}</div>', unsafe_allow_html=True)
        else:
            st.info("Chưa có gói thời gian nào.")

        st.markdown("### 🏋️ Gói theo buổi PT")
        sess_pkgs = [p for p in st.session_state.custom_packages if p["type"]=="session"]
        if sess_pkgs:
            cols2 = st.columns(min(len(sess_pkgs), 3))
            for i, pkg in enumerate(sess_pkgs):
                users_on_pkg2 = [m["customer_name"] for m in all_m if m.get("package_name","") == pkg["name"] and m["status"]=="active"]
                with cols2[i % 3]:
                    users_html2 = f'<div style="font-size:.75rem;color:var(--success);margin-top:.3rem">👥 {", ".join(users_on_pkg2)}</div>' if users_on_pkg2 else ""
                    st.markdown(f'<div class="card card-pt"><div style="font-family:\'Bebas Neue\';font-size:1.1rem;color:#FF6B35">{pkg["name"]}</div><div style="font-size:1.4rem;font-weight:800;margin:.3rem 0">{pkg["price"]:,.0f}<span style="font-size:.75rem"> đ</span></div><div style="font-size:.82rem;color:var(--subtext)">🏋️ {pkg["sessions"]} buổi</div>{users_html2}</div>', unsafe_allow_html=True)
        else:
            st.info("Chưa có gói buổi PT nào.")

    with tab2:
        c1, c2 = st.columns(2)
        with c1:
            pg_name = st.text_input("Tên gói *")
            pg_price = st.number_input("Giá (VNĐ)", 0, 99999999, 1000000, step=100000)
            pg_type = st.selectbox("Loại gói", ["⏱ Theo thời gian", "🏋️ Theo buổi PT"])
        with c2:
            if "thời gian" in pg_type:
                pg_dur = st.number_input("Thời hạn (ngày)", 1, 730, 30)
                pg_sess = None
            else:
                pg_sess = st.number_input("Số buổi", 1, 500, 10)
                pg_dur = None
            pg_note = st.text_area("Mô tả gói")

        if st.button("💾 Tạo gói", use_container_width=True):
            if pg_name:
                new_pkg = {"id": f"cp{len(st.session_state.custom_packages)+1}", "name": pg_name, "price": pg_price, "note": pg_note}
                if pg_sess is None:
                    new_pkg.update({"type":"time","duration":pg_dur})
                else:
                    new_pkg.update({"type":"session","sessions":pg_sess})
                st.session_state.custom_packages.append(new_pkg)
                st.success(f"✅ Đã tạo gói '{pg_name}'!"); st.rerun()
            else:
                st.error("Vui lòng nhập tên gói!")

    with tab3:
        st.markdown("### 🗑️ Xóa gói tập")
        if not st.session_state.custom_packages:
            st.info("Không có gói nào để xóa.")
        else:
            pkg_names = [p["name"] for p in st.session_state.custom_packages]
            del_pkg = st.selectbox("Chọn gói cần xóa:", ["-- Chọn gói --"] + pkg_names)
            if del_pkg != "-- Chọn gói --":
                pkg_obj = next((p for p in st.session_state.custom_packages if p["name"] == del_pkg), None)
                users_using = [m["customer_name"] for m in all_m if m.get("package_name","") == del_pkg and m["status"]=="active"]
                if users_using:
                    st.markdown(f'<div class="confirm-box"><div class="confirm-title">⚠️ Gói đang được sử dụng!</div><div style="font-size:.88rem;margin-top:.5rem">Gói <b>{del_pkg}</b> đang được dùng bởi: <b>{", ".join(users_using)}</b>.<br>Không thể xóa khi còn khách hàng đang sử dụng.</div></div>', unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="card" style="border-left-color:var(--danger)"><b>{del_pkg}</b> — Không có khách hàng đang sử dụng</div>', unsafe_allow_html=True)
                    if st.button(f"🗑️ Xóa gói '{del_pkg}'", use_container_width=True):
                        st.session_state.custom_packages = [p for p in st.session_state.custom_packages if p["name"] != del_pkg]
                        st.success(f"✅ Đã xóa gói '{del_pkg}'!"); st.rerun()

def page_reports():
    st.markdown("# 📈 Báo cáo & Thống kê"); st.markdown("---")
    all_m = st.session_state.memberships
    all_customers = st.session_state.get("customers_list", [])
    all_pts = st.session_state.get("pts_list", [])

    n_customers = len(all_customers) if all_customers else len({m["customer_id"] for m in all_m})
    n_pts = len(all_pts)
    n_time = sum(1 for m in all_m if m["type"]=="time" and m["status"]=="active")
    n_sess = sum(1 for m in all_m if m["type"]=="session" and m["status"]=="active")

    # Tổng quan
    c1,c2,c3,c4 = st.columns(4)
    for lbl,val,sub,col in [
        ("Tổng số khách hàng", n_customers, "thành viên", c1),
        ("Tổng số PT", n_pts, "huấn luyện viên", c2),
        ("Tổng gói PT", n_sess, "đang active", c3),
        ("Tổng gói thời gian", n_time, "đang active", c4),
    ]:
        with col: st.markdown(f'<div class="mbox"><div class="val">{val}</div><div class="lbl">{lbl}</div><div class="sub">{sub}</div></div>', unsafe_allow_html=True)

    st.markdown("---")

    # Bộ lọc thời gian
    st.markdown("### 🗓️ Lọc theo thời gian")
    filter_col1, filter_col2 = st.columns([2,3])
    with filter_col1:
        period = st.selectbox("Chọn kỳ báo cáo:", ["📅 Tuần này","📅 Tháng này","📅 Năm này","🗓️ Tùy chỉnh"])
    with filter_col2:
        today = date.today()
        if "Tuần" in period:
            start_filter = today - timedelta(days=today.weekday())
            end_filter = start_filter + timedelta(days=6)
        elif "Tháng" in period:
            start_filter = today.replace(day=1)
            end_filter = today
        elif "Năm" in period:
            start_filter = today.replace(month=1,day=1)
            end_filter = today
        else:
            fc1, fc2 = st.columns(2)
            with fc1: start_filter = st.date_input("Từ ngày", value=today.replace(day=1))
            with fc2: end_filter = st.date_input("Đến ngày", value=today)
        if "Tùy" not in period:
            st.markdown(f'<div style="font-size:.85rem;color:var(--subtext);padding:.5rem 0">📆 {fmt_date(str(start_filter))} → {fmt_date(str(end_filter))}</div>', unsafe_allow_html=True)

    # Lọc memberships theo thời gian
    def in_range(m):
        try:
            sd = date.fromisoformat(m.get("start_date",""))
            return start_filter <= sd <= end_filter
        except: return False

    filtered_m = [m for m in all_m if in_range(m)]
    f_time = [m for m in filtered_m if m["type"]=="time"]
    f_sess = [m for m in filtered_m if m["type"]=="session"]
    total_done = sum(m.get("sessions_done",0) for m in filtered_m if m["type"]=="session")

    st.markdown("#### 📊 Kết quả theo kỳ đã chọn")
    r1,r2,r3,r4 = st.columns(4)
    for lbl,val,col in [("Gói thời gian mới",len(f_time),r1),("Gói PT mới",len(f_sess),r2),("Tổng buổi đã tập",total_done,r3),("Tổng gói mới",len(filtered_m),r4)]:
        with col: st.markdown(f'<div class="mbox"><div class="val">{val}</div><div class="lbl">{lbl}</div></div>',unsafe_allow_html=True)

    st.markdown("---")
    # Cảnh báo
    ca, cb = st.columns(2)
    with ca:
        st.markdown("### 📦 Gói thời gian sắp hết (≤14 ngày)")
        found = False
        for m in all_m:
            if m["type"]=="time":
                dl = days_left(m.get("end_date",""))
                if dl is not None and 0<=dl<=14:
                    found=True; color="var(--danger)" if dl<=7 else "var(--warning)"
                    st.markdown(f'<div class="card card-warn"><b>{m["customer_name"]}</b><br><span style="font-size:.82rem;color:var(--subtext)">{m["package_name"]}</span><span style="float:right;color:{color};font-weight:700">Còn {dl} ngày</span></div>', unsafe_allow_html=True)
        if not found: st.info("Không có gói nào sắp hết hạn.")
    with cb:
        st.markdown("### 🏋️ Gói PT sắp hết buổi (≤5 buổi)")
        found2 = False
        for m in all_m:
            if m["type"]=="session":
                left = sessions_remaining(m)
                if 0<left<=5:
                    found2=True
                    st.markdown(f'<div class="card card-warn"><b>{m["customer_name"]}</b><br><span style="font-size:.82rem;color:var(--subtext)">{m["package_name"]}</span><span style="float:right;color:var(--warning);font-weight:700">Còn {left} buổi</span></div>', unsafe_allow_html=True)
        if not found2: st.info("Không có gói nào sắp hết buổi.")

    # Tải file Excel
    st.markdown("---")
    st.markdown("### 📥 Tải file tổng hợp")
    dl_col1, dl_col2 = st.columns([2,2])
    with dl_col1:
        if st.button("📊 Tải báo cáo Excel", use_container_width=True):
            try:
                import io, openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                wb = openpyxl.Workbook()
                ws = wb.active; ws.title = "Báo cáo tổng hợp"
                # Header
                ws.append(["STT","Khách hàng","Gói tập","Loại gói","PT","Ngày bắt đầu","Ngày kết thúc / Số buổi","Trạng thái"])
                for i, m in enumerate(all_m, 1):
                    if m["type"]=="time":
                        extra = m.get("end_date","")
                    else:
                        extra = f'{sessions_remaining(m)}/{m.get("sessions_total",0)} buổi còn lại'
                    ws.append([i, m.get("customer_name",""), m.get("package_name",""),
                                "Theo thời gian" if m["type"]=="time" else "Theo buổi PT",
                                m.get("pt_name","—"), m.get("start_date",""), extra, m.get("status","")])
                # Style header
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(fill_type="solid", fgColor="FF6B35")
                    cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = 20
                buf = io.BytesIO(); wb.save(buf); buf.seek(0)
                st.download_button("⬇️ Tải xuống file Excel", data=buf, file_name=f"baocao_fitpro_{today}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except ImportError:
                st.error("Cần cài openpyxl: pip install openpyxl")

def page_pts():
    st.markdown("# 🏋️ Quản lý Personal Trainer"); st.markdown("---")

    pt_list = st.session_state.get("pts_list", [])
    if not pt_list:
        pt_list = [
            {"name":"Nguyễn Văn A","email":"pt1@fitpro.vn","phone":"0934567890","specialty":"Cardio, Weight Loss","experience":"3 năm","gender":"Nam","created_at":"2024-01-01"},
            {"name":"Trần Thị B","email":"pt2@fitpro.vn","phone":"0945678901","specialty":"Yoga, Muscle","experience":"2 năm","gender":"Nữ","created_at":"2024-03-15"},
        ]

    def count_customers(pt_name):
        return len(set(m["customer_id"] for m in st.session_state.memberships if m.get("pt_name","") == pt_name or m.get("pt_name","") == f"PT {pt_name}"))

    pt_names = [pt.get("name","") for pt in pt_list]
    tab1, tab2 = st.tabs(["📋 Danh sách PT", "➕ Thêm PT mới"])

    with tab1:
        selected_pt = st.selectbox("🔍 Chọn PT để xem chi tiết:", ["-- Chọn PT --"] + pt_names)

        if selected_pt != "-- Chọn PT --":
            pt_info = next((pt for pt in pt_list if pt.get("name","") == selected_pt), None)
            pt_memberships = [m for m in st.session_state.memberships if m.get("pt_name","") == selected_pt or m.get("pt_name","") == f"PT {selected_pt}"]

            st.markdown(f"### 💪 {selected_pt}")
            n_cust = count_customers(selected_pt)

            # --- Thông tin cá nhân ---
            with st.expander("📋 Thông tin cá nhân", expanded=True):
                if pt_info:
                    c1, c2 = st.columns(2)
                    with c1:
                        st.markdown(f"""
                        <div class="card card-pt">
                            <div style="font-size:.9rem;line-height:1.9">
                                📱 <b>SĐT:</b> {pt_info.get("phone","—")}<br>
                                🚻 <b>Giới tính:</b> {pt_info.get("gender","—")}<br>
                                🎯 <b>Chuyên môn:</b> {pt_info.get("specialty","—")}<br>
                                ⏱ <b>Kinh nghiệm:</b> {pt_info.get("experience","—")}<br>
                                📅 <b>Ngày đăng ký:</b> {fmt_date(pt_info.get("created_at","")[:10] if pt_info.get("created_at") else "")}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    with c2:
                        st.markdown(f"""
                        <div class="card card-pt">
                            <div style="font-size:.9rem;line-height:1.9">
                                📧 <b>Email:</b> {pt_info.get("email","—")}<br>
                                🔑 <b>Mật khẩu:</b> {"•"*len(pt_info.get("password","")) if pt_info.get("password") else "—"}<br>
                                👥 <b>Số khách đang quản lý:</b> {n_cust}<br>
                                🏆 <b>Chứng chỉ:</b> {pt_info.get("certificates","—")}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

            # --- Khách hàng của PT ---
            if pt_memberships:
                st.markdown("#### 👥 Khách hàng đang quản lý")
                for m in pt_memberships:
                    render_membership_card(m, show_customer=True, show_pt=False)
            else:
                st.info("PT này chưa được phân công khách hàng nào.")

            # --- Chuyển PT ---
            st.markdown("---")
            with st.expander("🔄 Chuyển PT cho khách hàng", expanded=False):
                st.markdown("Chọn khách hàng và PT mới để chuyển:")
                other_pts = [pt.get("name","") for pt in pt_list if pt.get("name","") != selected_pt]
                if pt_memberships and other_pts:
                    transfer_customers = [m["customer_name"] for m in pt_memberships]
                    tr_cust = st.selectbox("Chọn khách hàng cần chuyển:", transfer_customers, key=f"tr_cust_{selected_pt}")
                    tr_pt = st.selectbox("Chuyển sang PT:", other_pts, key=f"tr_pt_{selected_pt}")
                    tr_new_pt_id = next((pt.get("id","") for pt in pt_list if pt.get("name","") == tr_pt), tr_pt)
                    if st.button("🔄 Xác nhận chuyển PT", key=f"tr_btn_{selected_pt}", use_container_width=True):
                        for m in st.session_state.memberships:
                            if m.get("customer_name","") == tr_cust and (m.get("pt_name","") == selected_pt or m.get("pt_name","") == f"PT {selected_pt}"):
                                m["pt_name"] = tr_pt
                                m["pt_id"] = tr_new_pt_id
                        st.success(f"✅ Đã chuyển {tr_cust} sang PT {tr_pt}!"); st.rerun()
                elif not other_pts:
                    st.warning("Không có PT khác để chuyển. Vui lòng thêm PT mới trước.")
                else:
                    st.info("PT này chưa có khách hàng nào.")

            # --- Xóa PT ---
            st.markdown("---")
            with st.expander("🗑️ Xóa PT này", expanded=False):
                active_customers_of_pt = [m["customer_name"] for m in pt_memberships]
                if active_customers_of_pt:
                    st.markdown(f'<div class="confirm-box"><div class="confirm-title">⚠️ PT này đang có {len(active_customers_of_pt)} khách hàng!</div><div style="font-size:.88rem;margin-top:.5rem">Bạn phải chuyển tất cả khách hàng sang PT khác trước khi xóa PT <b>{selected_pt}</b>.</div></div>', unsafe_allow_html=True)
                    other_pts2 = [pt.get("name","") for pt in pt_list if pt.get("name","") != selected_pt]
                    if other_pts2:
                        bulk_pt = st.selectbox("Chuyển tất cả khách sang PT:", other_pts2, key=f"bulk_pt_{selected_pt}")
                        bulk_pt_id = next((pt.get("id","") for pt in pt_list if pt.get("name","") == bulk_pt), bulk_pt)
                        confirm_del = st.text_input(f"Nhập tên PT '{selected_pt}' để xác nhận xóa sau khi chuyển:", key=f"del_pt_confirm_{selected_pt}")
                        if st.button("🗑️ Chuyển PT & Xóa", key=f"del_pt_btn_{selected_pt}", type="primary"):
                            if confirm_del == selected_pt:
                                for m in st.session_state.memberships:
                                    if m.get("pt_name","") == selected_pt or m.get("pt_name","") == f"PT {selected_pt}":
                                        m["pt_name"] = bulk_pt; m["pt_id"] = bulk_pt_id
                                st.session_state.pts_list = [pt for pt in st.session_state.pts_list if pt.get("name","") != selected_pt]
                                if db and pt_info and pt_info.get("id"):
                                    db.collection("users").document(pt_info["id"]).delete()
                                st.success(f"✅ Đã chuyển khách sang {bulk_pt} và xóa PT {selected_pt}!"); st.rerun()
                            else:
                                st.error("Tên xác nhận không khớp!")
                    else:
                        st.error("Không có PT nào khác! Thêm PT mới trước khi xóa.")
                else:
                    confirm_del2 = st.text_input(f"Nhập tên PT '{selected_pt}' để xác nhận xóa:", key=f"del_pt_confirm2_{selected_pt}")
                    if st.button("🗑️ Xác nhận xóa PT", key=f"del_pt_btn2_{selected_pt}", type="primary"):
                        if confirm_del2 == selected_pt:
                            st.session_state.pts_list = [pt for pt in st.session_state.pts_list if pt.get("name","") != selected_pt]
                            if db and pt_info and pt_info.get("id"):
                                db.collection("users").document(pt_info["id"]).delete()
                            st.success(f"✅ Đã xóa PT {selected_pt}!"); st.rerun()
                        else:
                            st.error("Tên xác nhận không khớp!")
        else:
            # Hiển thị danh sách tất cả PT
            st.markdown(f"#### 📋 Tổng số: {len(pt_list)} PT")
            for pt in pt_list:
                n_customers = count_customers(pt.get("name",""))
                st.markdown(f'''
                    <div class="card card-pt">
                        <div style="display:flex;justify-content:space-between">
                            <b>💪 {pt.get("name","")}</b>
                            <span class="badge b-active">Đang làm việc</span>
                        </div>
                        <div style="display:grid;grid-template-columns:repeat(3,1fr);gap:.3rem;margin-top:.5rem;font-size:.82rem;color:var(--subtext)">
                            <div>📧 {pt.get("email","")}</div>
                            <div>📱 {pt.get("phone","")}</div>
                            <div>⏱ {pt.get("experience","")}</div>
                        </div>
                        <div style="margin-top:.4rem;font-size:.83rem">🎯 {pt.get("specialty","")} &nbsp;|&nbsp; 👥 {n_customers} khách hàng</div>
                    </div>
                ''', unsafe_allow_html=True)

    with tab2:
        c1,c2=st.columns(2)
        with c1:
            p_name=st.text_input("Họ và tên PT *")
            p_email=st.text_input("Email *")
            p_phone=st.text_input("Số điện thoại")
            p_pwd=st.text_input("Mật khẩu *",type="password")
            p_gender=st.selectbox("Giới tính",["Nam","Nữ","Khác"])
        with c2:
            p_spec=st.text_input("Chuyên môn")
            p_exp=st.text_input("Kinh nghiệm")
            p_cert=st.text_area("Chứng chỉ")
        if st.button("💾 Thêm PT",use_container_width=True):
            if p_name and p_email and p_pwd:
                if db:
                    doc_ref = db.collection("users").add({
                        "name":p_name,"email":p_email,"phone":p_phone,"password":p_pwd,
                        "role":"pt","specialty":p_spec,"experience":p_exp,"certificates":p_cert,
                        "gender":p_gender,"created_at":datetime.now().isoformat()
                    })
                st.session_state.pts_list.append({
                    "name":p_name,"email":p_email,"phone":p_phone,"password":p_pwd,
                    "role":"pt","specialty":p_spec,"experience":p_exp,"gender":p_gender,
                    "created_at":datetime.now().isoformat()
                })
                st.success(f"✅ Đã thêm PT {p_name}!"); st.rerun()
            else: st.error("Điền đầy đủ thông tin!")

def main():
    if not st.session_state.logged_in: page_login(); return
    render_sidebar()
    routes={"dashboard":page_dashboard,"my_memberships":page_my_memberships,"confirm_session":page_confirm_session,"my_sessions":page_my_sessions,"progress":page_progress,"my_customers":page_my_customers,"record_session":page_record_session,"pt_sessions":page_pt_sessions,"customers":page_customers_owner,"packages":page_packages,"reports":page_reports,"pts":page_pts}
    routes.get(st.session_state.page,page_dashboard)()

if __name__=="__main__": main()
